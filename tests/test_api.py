from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app
from app.store import store


@pytest.fixture(autouse=True)
def reset_store(tmp_path, monkeypatch):
    """Keep API tests isolated while exercising the real in-memory store."""
    store.reset()
    monkeypatch.setattr(store, "upload_root", tmp_path / "uploads")
    monkeypatch.setenv("APP_ACCESS_PASSWORD", "test-password")


@pytest.fixture
def client() -> TestClient:
    return TestClient(app, headers={"X-Access-Password": "test-password"})


def workbook_bytes() -> bytes:
    workbook = Workbook()
    workbook.active.append(["项目", "金额"])
    workbook.active.append(["测试", 12])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def upload(client: TestClient, *files: tuple[str, bytes, str]) -> dict:
    response = client.post(
        "/documents",
        files=[("files", item) for item in files],
    )
    assert response.status_code == 200, response.text
    return response.json()


def test_upload_process_review_and_export_only_confirmed_documents(client: TestClient) -> None:
    created = upload(
        client,
        ("records.csv", b"Name,City\nAlice,Shanghai\n", "text/csv"),
        (
            "ledger.xlsx",
            workbook_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    )
    batch_id = created["batch_id"]
    assert len(created["documents"]) == 2

    processed = client.post(f"/batches/{batch_id}/process")
    assert processed.status_code == 200
    assert processed.json()["progress"] == {"uploaded": 0, "processing": 0, "processed": 2, "failed": 0}

    document_id = created["documents"][0]["id"]
    reviewed = client.patch(
        f"/documents/{document_id}/review",
        json={
            "classification": "人工分类",
            "summary": "人工摘要",
            "key_fields": {"负责人": "Alice"},
            "confidence": 0.9,
            "evidence": [{"location": "row 2", "text": "Alice"}],
            "review_status": "confirmed",
            "reviewer_notes": "复核完成",
        },
    )
    assert reviewed.status_code == 200
    assert reviewed.json()["summary"] == "人工摘要"

    export = client.get(f"/batches/{batch_id}/export")
    assert export.status_code == 200
    assert export.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    sheet = load_workbook(BytesIO(export.content), read_only=True).active
    rows = list(sheet.values)
    assert rows[0] == ("来源文件", "资料类型", "分类", "摘要", "关键字段", "置信度", "依据", "审核状态", "人工备注")
    assert len(rows) == 2
    assert rows[1][0] == "records.csv"
    assert rows[1][2:4] == ("人工分类", "人工摘要")
    assert rows[1][7] == "confirmed"


def test_invalid_upload_does_not_block_valid_document(client: TestClient) -> None:
    created = upload(
        client,
        ("safe.csv", b"name\nAlice\n", "text/csv"),
        ("notes.txt", b"must be refused", "text/plain"),
    )

    assert [document["filename"] for document in created["documents"]] == ["safe.csv"]
    assert created["errors"] == [{"filename": "notes.txt", "error": "不支持的文件类型"}]

    batch = client.get(f"/batches/{created['batch_id']}")
    assert batch.status_code == 200
    assert batch.json()["progress"]["uploaded"] == 1


def test_processing_bad_file_continues_and_reprocess_retries(client: TestClient) -> None:
    created = upload(
        client,
        ("valid.csv", b"item\nkept\n", "text/csv"),
        ("bad.xlsx", b"not an excel workbook", "application/vnd.ms-excel"),
    )
    batch_id = created["batch_id"]

    first = client.post(f"/batches/{batch_id}/process")
    assert first.status_code == 200
    state = client.get(f"/batches/{batch_id}").json()
    statuses = {document["filename"]: document["status"] for document in state["documents"]}
    assert statuses == {"valid.csv": "processed", "bad.xlsx": "failed"}
    assert state["progress"] == {"uploaded": 0, "processing": 0, "processed": 1, "failed": 1}

    second = client.post(f"/batches/{batch_id}/process")
    assert second.status_code == 200
    state_after_retry = client.get(f"/batches/{batch_id}").json()
    assert {document["filename"]: document["status"] for document in state_after_retry["documents"]} == statuses

    valid_id = next(document["id"] for document in state["documents"] if document["filename"] == "valid.csv")
    marked = client.patch(f"/documents/{valid_id}/review", json={"review_status": "reprocess"})
    assert marked.status_code == 200
    client.post(f"/batches/{batch_id}/process")
    assert client.get(f"/batches/{batch_id}").json()["documents"][0]["status"] == "processed"


def test_batch_status_is_safe_and_unknown_resources_are_404(client: TestClient) -> None:
    secret = "联系电话 13800138000"
    created = upload(client, ("private.csv", f"note\n{secret}\n".encode(), "text/csv"))
    batch_id = created["batch_id"]
    client.post(f"/batches/{batch_id}/process")

    response = client.get(f"/batches/{batch_id}")
    assert response.status_code == 200
    serialized = response.text
    assert "raw_text" not in serialized
    assert "redacted_text" not in serialized
    assert secret not in serialized
    assert "[REDACTED_PHONE]" not in serialized

    assert client.get("/batches/missing").status_code == 404
    assert client.patch("/documents/missing/review", json={"review_status": "confirmed"}).status_code == 404


def test_empty_confirmed_export_has_headers_only(client: TestClient) -> None:
    created = upload(client, ("waiting.csv", b"item\nno approval\n", "text/csv"))
    export = client.get(f"/batches/{created['batch_id']}/export")
    sheet = load_workbook(BytesIO(export.content), read_only=True).active
    assert list(sheet.values) == [
        ("来源文件", "资料类型", "分类", "摘要", "关键字段", "置信度", "依据", "审核状态", "人工备注")
    ]


def test_index_exposes_archive_workbench_landmarks(client: TestClient) -> None:
    response = client.get("/")

    assert response.status_code == 200
    assert 'class="workbench"' in response.text
    assert 'data-ui="archive-workbench"' in response.text
    assert 'id="files"' in response.text
    assert 'id="rows"' in response.text
    assert 'id="status"' in response.text


def test_api_rejects_requests_without_the_shared_password() -> None:
    with TestClient(app) as unauthenticated_client:
        response = unauthenticated_client.get("/batches/missing")

    assert response.status_code == 401
    assert response.json()["detail"] == "Access password required"


def test_access_verify_accepts_only_the_shared_password(client: TestClient) -> None:
    accepted = client.get("/access/verify")

    assert accepted.status_code == 204

    with TestClient(app) as unauthenticated_client:
        rejected = unauthenticated_client.get("/access/verify")

    assert rejected.status_code == 401


def test_index_includes_the_shared_access_gate(client: TestClient) -> None:
    response = client.get("/")

    assert response.status_code == 200
    assert 'id="access-gate"' in response.text
    assert "X-Access-Password" in response.text


def test_access_gate_loads_the_external_browser_script(client: TestClient) -> None:
    response = client.get("/")

    assert '<script src="/config.js"></script>' in response.text
    assert '<script src="/app.js"></script>' in response.text
    script = (Path(__file__).parents[1] / "app" / "templates" / "app.js").read_text(
        encoding="utf-8"
    )
    assert "form.addEventListener('submit', unlock)" in script


def test_index_keeps_access_password_in_memory_when_browser_storage_is_unavailable(client: TestClient) -> None:
    response = client.get("/")

    assert response.status_code == 200
    assert "let transientAccessPassword=''" in response.text
    assert "try{localStorage.setItem" in response.text


def test_index_binds_access_gate_with_a_dom_submit_listener(client: TestClient) -> None:
    response = client.get("/")

    assert response.status_code == 200
    script = (Path(__file__).parents[1] / "app" / "templates" / "app.js").read_text(
        encoding="utf-8"
    )
    assert "form.addEventListener('submit', unlock)" in script
